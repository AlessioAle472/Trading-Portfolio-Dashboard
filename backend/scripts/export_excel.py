import sys
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# ─── PALETTE COLORI ──────────────────────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # blu notte
C_HEADER_FG   = "E2C27D"   # oro
C_TITLE_BG    = "16213E"   # blu scuro per titolo foglio
C_TITLE_FG    = "FFFFFF"
C_ROW_ODD     = "F7F9FC"   # grigio chiarissimo
C_ROW_EVEN    = "FFFFFF"
C_BORDER      = "B0BEC5"
C_LOTTI       = "E8F5E9"   # verde chiaro per lotti
C_ALERT_COL   = "FFF9C4"   # giallo per colonna stato allerta

def make_border(color=C_BORDER):
    thin = Side(style="thin", color=color)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_font():
    return Font(name="Calibri", bold=True, color=C_HEADER_FG, size=11)

def title_font():
    return Font(name="Calibri", bold=True, color=C_TITLE_FG, size=13)

def data_font(bold=False):
    return Font(name="Calibri", bold=bold, size=10)

HEADERS = [
    "Mercato",
    "Nome Strategia / Portafoglio",
    "TimeFrame",
    "MagicNumber",
    "In Reale? / Come sta andando? / Note",
    "Lotti Attuali",
    "Monte Carlo DD (%)",
    "DD Reale Live (%)",
    "Stato Allerta",
]

# Column widths
COL_WIDTHS = [14, 36, 12, 32, 52, 14, 20, 20, 16]

COL_LOTTI       = 6   # F
COL_MC_DD       = 7   # G
COL_REAL_DD     = 8   # H
COL_ALERT       = 9   # I

def export_excel(json_path, output_path):
    if not os.path.exists(json_path):
        print(f"Error: JSON file not found at {json_path}")
        sys.exit(1)
        
    with open(json_path, 'r', encoding='utf-8') as f:
        sheet_data = json.load(f)
        
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet
    
    # Get sheets dynamically from the keys of the JSON file
    valid_sheets = list(sheet_data.keys())
    
    for sheet_name in valid_sheets:
        rows = sheet_data[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        
        # ── row 1: title ─────────────────────────────────────────────
        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
        title_cell = ws.cell(row=1, column=1, value=f"📊  PORTAFOGLIO – {sheet_name.upper()}")
        title_cell.font      = title_font()
        title_cell.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # ── row 2: headers ───────────────────────────────────────────
        ws.row_dimensions[2].height = 22
        for col_idx, hdr in enumerate(HEADERS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=hdr)
            cell.font      = header_font()
            cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = make_border()
            
        # ── data rows ────────────────────────────────────────────────
        for row_idx, row in enumerate(rows, start=3):
            is_odd = (row_idx % 2 == 1)
            row_bg = C_ROW_ODD if is_odd else C_ROW_EVEN
            
            # col A - Mercato
            c = ws.cell(row=row_idx, column=1, value=row.get("mercato", ""))
            c.font      = data_font(bold=True)
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = make_border()
            
            # col B - Nome
            c = ws.cell(row=row_idx, column=2, value=row.get("nome", ""))
            c.font      = data_font()
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = make_border()
            
            # col C - TimeFrame
            c = ws.cell(row=row_idx, column=3, value=row.get("tf", ""))
            c.font      = data_font()
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = make_border()
            
            # col D - MagicNumber
            c = ws.cell(row=row_idx, column=4, value=row.get("magic", ""))
            c.font      = data_font()
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = make_border()
            
            # col E - Note
            c = ws.cell(row=row_idx, column=5, value=row.get("note", ""))
            c.font      = data_font()
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = make_border()
            
            # col F - Lotti Attuali
            lotti_val = row.get("lotti", 0.01)
            try:
                lotti_val = float(lotti_val)
            except (ValueError, TypeError):
                lotti_val = 0.01
            c = ws.cell(row=row_idx, column=COL_LOTTI, value=lotti_val)
            c.font      = data_font(bold=True)
            c.fill      = PatternFill("solid", fgColor=C_LOTTI)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00"
            c.border    = make_border()
            
            # col G - Monte Carlo DD (%)
            mc_val = row.get("mc_dd")
            if mc_val is not None and mc_val != "":
                try:
                    mc_val = float(mc_val)
                except ValueError:
                    pass
            else:
                mc_val = None
            c = ws.cell(row=row_idx, column=COL_MC_DD, value=mc_val)
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = '0.00"%"'
            c.border    = make_border()
            
            # col H - DD Reale Live (%)
            real_val = row.get("real_dd")
            if real_val is not None and real_val != "":
                try:
                    real_val = float(real_val)
                except ValueError:
                    pass
            else:
                real_val = None
            c = ws.cell(row=row_idx, column=COL_REAL_DD, value=real_val)
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = '0.00"%"'
            c.border    = make_border()
            
            # col I - Stato Allerta (Excel Formula)
            g_ref = f"G{row_idx}"
            h_ref = f"H{row_idx}"
            formula = (
                f'=IF(OR({g_ref}="",{h_ref}=""),"OK",'
                f'IF({h_ref}>{g_ref},"CRITICO",'
                f'IF({h_ref}>{g_ref}*0.8,"ATTENZIONE","OK")))'
            )
            c = ws.cell(row=row_idx, column=COL_ALERT, value=formula)
            c.font      = data_font(bold=True)
            c.fill      = PatternFill("solid", fgColor=C_ALERT_COL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = make_border()
            
            ws.row_dimensions[row_idx].height = 20
            
        # ── conditional formatting for alert column ──────────────────────────
        last_data_row = 2 + len(rows)
        if last_data_row >= 3:
            alert_range = f"I3:I{last_data_row}"
            
            # CRITICO -> Red
            ws.conditional_formatting.add(
                alert_range,
                FormulaRule(
                    formula=[f'I3="CRITICO"'],
                    fill=PatternFill("solid", fgColor="FF4444"),
                    font=Font(bold=True, color="FFFFFF"),
                )
            )
            # ATTENZIONE -> Orange
            ws.conditional_formatting.add(
                alert_range,
                FormulaRule(
                    formula=[f'I3="ATTENZIONE"'],
                    fill=PatternFill("solid", fgColor="FFA500"),
                    font=Font(bold=True, color="FFFFFF"),
                )
            )
            # OK -> Green
            ws.conditional_formatting.add(
                alert_range,
                FormulaRule(
                    formula=[f'I3="OK"'],
                    fill=PatternFill("solid", fgColor="4CAF50"),
                    font=Font(bold=True, color="FFFFFF"),
                )
            )
            
        # ── widths ───────────────────────────────────────────────────────────
        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            
        # ── freeze panes ─────────────────────────────────────────────────────
        ws.freeze_panes = "A3"
        
        # ── auto filters ─────────────────────────────────────────────────────
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"
        
    wb.save(output_path)
    print(f"Excel successfully generated at {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 export_excel.py <json_path> <output_path>")
        sys.exit(1)
    export_excel(sys.argv[1], sys.argv[2])
